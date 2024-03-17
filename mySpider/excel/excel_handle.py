from openpyxl import load_workbook

from mySpider.db_dao.db_handle import get_collection
from mySpider.excel.excel_read import get_head_idx
from mySpider.items import ProductionItem
from mySpider.utils.cyberebeeUtils import handle_feature, handle_bullet_points
from mySpider.utils.myUtils import hash_str

base_path = "C:/Users/cainiao/Documents/爬虫/天猫商店/"

file_path = base_path + "cyberebee02.xlsx"
wb = load_workbook(filename=file_path, read_only=False, keep_vba=True)

ws = wb.active
header = {}
item_list = []

for r_num, row in enumerate(ws.iter_rows()):
    item = ProductionItem()
    for index, cell in enumerate(row, start=1):
        if r_num == 0:
            header[index] = cell.value
        else:
            if header[index] in item.fields:
                item[header[index]] = cell.value
            else:
                pass
    if r_num > 0 and item is not None:
        item_list.append(item)

# print("header=", header)
# print("item_list=", item_list)
# 数据处理
for item in item_list:
    item['manufacturer'] = 'IKEN'
    item['brand_name'] = 'IKEN'
    item['update_delete'] = 'Update'
    item['quantity'] = 0
    if item['standard_price']:
        item['standard_price'] = item['standard_price'].replace('$', '')
    item['source_item_id'] = hash_str(item['source_item_url'])
    item['source_item_url_hash'] = hash_str(item['source_item_url'])
    item['item_name'] = item['source_item_name']

    bullet_points = handle_feature(item['features'])
    if 'Features:' in bullet_points:
        # print("bullet_points=", bullet_points['Features:'])
        feature = bullet_points['Features:']
        try:
            points = handle_bullet_points(feature)
            # print("bullet_points=", points)
            for index, point in enumerate(points):
                num = index + 1
                if point and isinstance(point, str):
                    item['bullet_point' + str(num)] = point
        except Exception as e:
            print("feature=", feature)
            print(e)

# print(item_list[0])

# 数据落库
col = get_collection("demo", "spider1")
for item in item_list:
    result = col.find_one({'source_item_id': item['source_item_id']})
    if result:
        # 更新MongoDB数据
        # print("更新MongoDB数据2")
        col.update_one({'source_item_id': item['source_item_id']},
                       {'$set': dict(item)})
    else:
        col.insert_one(dict(item))

# 数据导出
filename = "E:/spider_data_export/Template_file.xlsx"
wb_export = load_workbook(filename=filename, read_only=False, keep_vba=True)
sheet = wb_export.active

documents = list(col.find())
for row_num, row_data in enumerate(documents, start=1):
    for key, value in row_data.items():
        head_idx = get_head_idx(key)
        if head_idx is None:
            continue
        if isinstance(value, dict):
            continue
        try:
            value = str(value)
            sheet.cell(row=row_num + 3, column=head_idx, value=value)
        except Exception as e:
            print(e)
            print("key=", key)
    # print(row_data)

wb_export.save(filename=filename)
