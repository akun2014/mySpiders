from marshmallow import Schema, fields
from openpyxl import load_workbook

"""
处理模板文件的标头，找到字段对应的index
"""


class Header(Schema):
    column_name = fields.Str()
    idx = fields.Integer()


wb = load_workbook(filename="../bizdata/template_file.xlsx", read_only=False, keep_vba=True)

ws = wb["Template"]
header = []

for row in ws.iter_rows(min_row=3, max_col=ws.max_column, max_row=3):
    for index, cell in enumerate(row):
        head = Header()
        head.column_name = cell.value
        head.idx = index
        header.append(head)

# json格式化
h = Header()
json_string = Header().dumps(header, many=True)
print(json_string)

# 写入文件
f = open("../bizdata/metadata.json")
f.write(json_string)
